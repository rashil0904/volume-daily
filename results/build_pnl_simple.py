#!/usr/bin/env python3
"""
build_pnl_simple.py -- generates strategy_pnl_simple.xlsx, a minimal 4-sheet
P&L tracker for a live NSE intraday strategy that trades both long and short
positions.

Sheet order: Total PnL, Trade Log, Day Wise PnL, Position Type Stats.
Every computed cell is a formula string -- openpyxl never pre-computes a
value in Python. Formulas avoid XLOOKUP/XMATCH/SORT/FILTER/UNIQUE/SEQUENCE
for LibreOffice/Google Sheets compatibility; MAXIFS/MINIFS are written as
_xlfn.MAXIFS/_xlfn.MINIFS so they don't show #NAME? in older Excel/LO
formula-function tables.

Usage:
    python3 build_pnl_simple.py
"""

import json
import re
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.line import LineProperties

RESULTS_DIR = Path(__file__).resolve().parent
OUT_PATH    = RESULTS_DIR / "strategy_pnl_simple.xlsx"
DHAN_POSITIONS_PATH = RESULTS_DIR / "positions_dhan.json"

# Trade Log is auto-synced from Dhan's real fills starting this date -- entries
# before it (from before this workbook existed) are out of scope, not missing data.
PNL_START_DATE = "2026-08-19"

# ── Number formats ──────────────────────────────────────────────────────────
INR   = '₹#,##0;(₹#,##0);"-"'
PCT   = '0.00%;(0.00%);"-"'
PCT1  = '0.0%;(0.0%);"-"'
NUM   = '#,##0;(#,##0);"-"'
PRICE = '#,##0.00'
DATE  = 'yyyy-mm-dd'
XMULT = '0.00"x"'

# ── Colors / fonts / fills ──────────────────────────────────────────────────
FONT_NAME = "Arial"

HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill(start_color="FF1F3864", end_color="FF1F3864", fill_type="solid")

INPUT_FONT = Font(name=FONT_NAME, color="FF0000FF")
INPUT_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

FORMULA_FONT = Font(name=FONT_NAME, color="FF000000")
LABEL_FONT   = Font(name=FONT_NAME, color="FF000000", bold=True)
TOTAL_FONT   = Font(name=FONT_NAME, color="FF000000", bold=True)

EXAMPLE_FONT = Font(name=FONT_NAME, italic=True, color="FF0000FF")
EXAMPLE_FILL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")

TITLE_FONT    = Font(name=FONT_NAME, bold=True, size=14, color="FF1F3864")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="FF808080")

RED_FONT      = Font(name=FONT_NAME, color="FFFF0000")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center")


# ── Styling helpers ──────────────────────────────────────────────────────────

def set_title_subtitle(ws, title: str, subtitle: str, ncols: int) -> None:
    last_col = get_column_letter(ncols)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = title
    c.font = TITLE_FONT
    c.alignment = LEFT
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = subtitle
    c.font = SUBTITLE_FONT
    c.alignment = LEFT
    ws.row_dimensions[2].height = 16


def style_header_row(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def style_input(cell, number_format: str | None = None) -> None:
    cell.font = INPUT_FONT
    cell.fill = INPUT_FILL
    if number_format:
        cell.number_format = number_format


def style_formula(cell, number_format: str | None = None) -> None:
    cell.font = FORMULA_FONT
    if number_format:
        cell.number_format = number_format


def style_example(cell, number_format: str | None = None) -> None:
    cell.font = EXAMPLE_FONT
    cell.fill = EXAMPLE_FILL
    if number_format:
        cell.number_format = number_format


# ── Sheet 2: Trade Log ───────────────────────────────────────────────────────

TL_FIRST_ROW = 4
TL_LAST_ROW  = 503
TL_HEADERS   = ["Trade ID", "Symbol", "Position", "Entry Date", "Entry Price",
                "Qty", "Exit Date", "Exit Price", "Gross P&L", "Costs",
                "Net P&L", "Net Return %", "Status"]


def _extract_exit(position: dict) -> tuple[str | None, float | None]:
    """Finds whichever exit_price_<stage>/exit_timestamp_<stage> pair is
    present on a position record -- stage suffix varies by exit path (925,
    1159 for a long; 239 for a mirrored short's square-off). Returns
    (iso_date_str, price), or (None, None) if the position is still open
    (no exit fields yet)."""
    for key, value in position.items():
        m = re.fullmatch(r"exit_price_(\w+)", key)
        if m and value is not None:
            ts = position.get(f"exit_timestamp_{m.group(1)}")
            if ts:
                return ts[:10], float(value)
    return None, None


def _position_to_trade_row(position: dict) -> dict | None:
    """Maps one positions_dhan.json record (long or mirrored short) to a
    Trade Log row dict. Costs has no source in position JSON -- it's never
    populated here, only ever carried forward from a prior workbook (see
    _load_existing_costs) or left for the user to fill in by hand."""
    symbol = position.get("symbol")
    if not symbol:
        return None

    is_short = position.get("direction") == "short"
    entry_price = position.get("entry_price") if is_short else position.get("actual_fill_price")
    qty         = position.get("quantity")     if is_short else position.get("actual_fill_quantity")
    if entry_price is None or qty is None:
        return None

    entry_date = position.get("entry_date") or (position.get("entry_timestamp") or "")[:10]
    if not entry_date or entry_date < PNL_START_DATE:
        return None

    exit_date, exit_price = _extract_exit(position)
    entry_order_id = position.get("entry_order_id") or ""
    trade_id = f"DHAN-{entry_order_id}" if entry_order_id else f"DHAN-{symbol}-{entry_date}"

    return {
        "id": trade_id, "symbol": symbol, "position": "SHORT" if is_short else "LONG",
        "entry_date": date.fromisoformat(entry_date), "entry_price": float(entry_price), "qty": int(qty),
        "exit_date": date.fromisoformat(exit_date) if exit_date else None,
        "exit_price": exit_price,
    }


def load_trades_from_positions() -> list[dict]:
    """Reads results/positions_dhan.json and returns one Trade Log row dict
    per position dated PNL_START_DATE or later (long or short) -- open
    positions come back with exit_date/exit_price=None, matching Trade
    Log's own Open/Closed status formula. Returns [] if the file is
    missing/unreadable/empty, in which case build_trade_log falls back to
    the two worked examples."""
    if not DHAN_POSITIONS_PATH.exists():
        return []
    try:
        positions = json.loads(DHAN_POSITIONS_PATH.read_text())
    except (json.JSONDecodeError, OSError):
        return []

    trades = [row for pos in positions if (row := _position_to_trade_row(pos))]
    trades.sort(key=lambda t: (t["entry_date"], t["symbol"]))
    return trades


def _load_existing_costs(path: Path) -> dict[str, float]:
    """Reads a previously-generated workbook's Trade Log Costs column, keyed
    by Trade ID, so re-running the auto-sync doesn't wipe out Costs the user
    has already typed in by hand -- everything else in a synced row comes
    fresh from position JSON on every run, but Costs has no JSON source, it
    only ever lives inside the workbook itself."""
    if not path.exists():
        return {}
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb["Trade Log"]
    except Exception:
        return {}
    costs = {}
    for r in range(TL_FIRST_ROW, TL_LAST_ROW + 1):
        trade_id = ws[f"A{r}"].value
        cost_val = ws[f"J{r}"].value
        if trade_id and isinstance(cost_val, (int, float)):
            costs[str(trade_id)] = cost_val
    return costs


def build_trade_log(ws, trades: list[dict] | None = None,
                     existing_costs: dict[str, float] | None = None) -> None:
    """trades=None (or empty) falls back to the two worked examples -- same
    as the original standalone builder. trades=[...] (from
    load_trades_from_positions) auto-syncs real Dhan fills into rows 4+
    instead; existing_costs (from _load_existing_costs) carries forward any
    Costs the user already typed in for a Trade ID that's still present."""
    existing_costs = existing_costs or {}

    set_title_subtitle(
        ws, "Trade Log",
        "One row per position -- long and short trades are logged separately. "
        "Fill columns A-H and J; formulas compute I, K, L, M.",
        len(TL_HEADERS),
    )

    for col, header in enumerate(TL_HEADERS, start=1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(TL_HEADERS))

    widths = [26, 14, 10, 12, 12, 9, 12, 12, 12, 10, 12, 12, 10]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    manual_fmts = {"D": DATE, "E": PRICE, "F": NUM, "G": DATE, "H": PRICE, "J": INR}
    synced_rows: dict[int, dict] = {}

    if trades:
        # Auto-synced from real Dhan fills -- still input-styled (blue/yellow)
        # since Costs (J) stays user-editable and everything regenerates from
        # position JSON on every run anyway.
        for offset, trade in enumerate(trades[: TL_LAST_ROW - TL_FIRST_ROW + 1]):
            r = TL_FIRST_ROW + offset
            synced_rows[r] = trade
            row_vals = {
                "A": trade["id"], "B": trade["symbol"], "C": trade["position"],
                "D": trade["entry_date"], "E": trade["entry_price"], "F": trade["qty"],
                "G": trade["exit_date"] or "", "H": trade["exit_price"] if trade["exit_price"] is not None else "",
                "J": existing_costs.get(trade["id"], 0),
            }
            for col_letter, val in row_vals.items():
                cell = ws[f"{col_letter}{r}"]
                cell.value = val
                style_input(cell, manual_fmts.get(col_letter))
    else:
        # No real trades yet -- two worked examples (rows 4-5), delete before real use.
        examples = [
            {
                "id": "EX-LONG-1 (EXAMPLE — delete before real use)",
                "symbol": "RELIANCE", "position": "LONG",
                "entry_date": date(2026, 7, 31), "entry_price": 2500.00, "qty": 100,
                "exit_date": date(2026, 8, 1), "exit_price": 2550.00, "costs": 150,
            },
            {
                "id": "EX-SHORT-1 (EXAMPLE — delete before real use)",
                "symbol": "TCS", "position": "SHORT",
                "entry_date": date(2026, 7, 31), "entry_price": 3600.00, "qty": 50,
                "exit_date": date(2026, 8, 1), "exit_price": 3550.00, "costs": 100,
            },
        ]
        for offset, ex in enumerate(examples):
            r = TL_FIRST_ROW + offset
            row_vals = {
                "A": ex["id"], "B": ex["symbol"], "C": ex["position"],
                "D": ex["entry_date"], "E": ex["entry_price"], "F": ex["qty"],
                "G": ex["exit_date"], "H": ex["exit_price"], "J": ex["costs"],
            }
            for col_letter, val in row_vals.items():
                cell = ws[f"{col_letter}{r}"]
                cell.value = val
                style_example(cell, manual_fmts.get(col_letter))

    # Manual input styling (for every row not already synced/example above) + formulas, rows 4-503.
    for r in range(TL_FIRST_ROW, TL_LAST_ROW + 1):
        if r not in synced_rows and not (not trades and r in (4, 5)):
            for col_letter in ("A", "B", "C", "D", "E", "F", "G", "H", "J"):
                cell = ws[f"{col_letter}{r}"]
                style_input(cell, manual_fmts.get(col_letter))

        i_formula = (f'=IF(OR($B{r}="",$H{r}=""),"",'
                     f'IF($C{r}="LONG",($H{r}-$E{r})*$F{r},($E{r}-$H{r})*$F{r}))')
        k_formula = f'=IF(OR($B{r}="",$I{r}=""),"",$I{r}-IF($J{r}="",0,$J{r}))'
        l_formula = f'=IF(OR($B{r}="",$K{r}="",$E{r}=0,$F{r}=0),"",$K{r}/($E{r}*$F{r}))'
        m_formula = f'=IF($B{r}="","",IF($H{r}="","Open","Closed"))'

        cell_i = ws[f"I{r}"]; cell_i.value = i_formula; style_formula(cell_i, INR)
        cell_k = ws[f"K{r}"]; cell_k.value = k_formula; style_formula(cell_k, INR)
        cell_l = ws[f"L{r}"]; cell_l.value = l_formula; style_formula(cell_l, PCT)
        cell_m = ws[f"M{r}"]; cell_m.value = m_formula; style_formula(cell_m)

    # Data validation: Position dropdown.
    dv = DataValidation(type="list", formula1='"LONG,SHORT"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"C{TL_FIRST_ROW}:C{TL_LAST_ROW}")

    # Conditional formatting: Net P&L < 0 -> red font.
    ws.conditional_formatting.add(
        f"K{TL_FIRST_ROW}:K{TL_LAST_ROW}",
        CellIsRule(operator="lessThan", formula=["0"], font=RED_FONT),
    )

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:M{TL_LAST_ROW}"


# ── Sheet 3: Day Wise PnL ────────────────────────────────────────────────────

DW_FIRST_ROW = 4
DW_LAST_ROW  = 403
DW_HEADERS   = ["Date", "Long Gross P&L", "Short Gross P&L", "Total Gross P&L",
                "Long Net P&L", "Short Net P&L", "Total Net P&L",
                "Cumulative Net P&L"]


def build_day_wise(ws) -> None:
    set_title_subtitle(
        ws, "Day Wise PnL",
        "One row per calendar date, attributed by Exit Date -- continues forward "
        "automatically from the first date in A4.",
        len(DW_HEADERS),
    )

    for col, header in enumerate(DW_HEADERS, start=1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(DW_HEADERS))

    widths = [14, 16, 16, 16, 16, 16, 16, 18]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # First date is a manual anchor -- matches the two example trades' exit date.
    a4 = ws["A4"]
    a4.value = date(2026, 8, 1)
    style_input(a4, DATE)

    tl = "'Trade Log'"
    for r in range(DW_FIRST_ROW, DW_LAST_ROW + 1):
        if r > DW_FIRST_ROW:
            a_formula = f'=IF(A{r-1}="","",A{r-1}+1)'
            cell_a = ws[f"A{r}"]; cell_a.value = a_formula; style_formula(cell_a, DATE)

        g = f'IF($A{r}="","",'

        b_formula = (f'={g}SUMIFS({tl}!$I$4:$I$503,{tl}!$G$4:$G$503,$A{r},'
                     f'{tl}!$C$4:$C$503,"LONG"))')
        c_formula = (f'={g}SUMIFS({tl}!$I$4:$I$503,{tl}!$G$4:$G$503,$A{r},'
                     f'{tl}!$C$4:$C$503,"SHORT"))')
        d_formula = f'=IF($A{r}="","",$B{r}+$C{r})'
        e_formula = (f'={g}SUMIFS({tl}!$K$4:$K$503,{tl}!$G$4:$G$503,$A{r},'
                     f'{tl}!$C$4:$C$503,"LONG"))')
        f_formula = (f'={g}SUMIFS({tl}!$K$4:$K$503,{tl}!$G$4:$G$503,$A{r},'
                     f'{tl}!$C$4:$C$503,"SHORT"))')
        g_formula = f'=IF($A{r}="","",$E{r}+$F{r})'
        prev = "0" if r == DW_FIRST_ROW else f"$H{r-1}"
        h_formula = f'={g}{prev}+IF($G{r}="",0,$G{r}))'

        for col_letter, formula in (
            ("B", b_formula), ("C", c_formula), ("D", d_formula),
            ("E", e_formula), ("F", f_formula), ("G", g_formula),
            ("H", h_formula),
        ):
            cell = ws[f"{col_letter}{r}"]
            cell.value = formula
            style_formula(cell, INR)

    ws.conditional_formatting.add(
        f"G{DW_FIRST_ROW}:G{DW_LAST_ROW}",
        CellIsRule(operator="lessThan", formula=["0"], font=RED_FONT),
    )

    ws.freeze_panes = "B4"


# ── Sheet 4: Position Type Stats ─────────────────────────────────────────────

PS_HEADERS = ["Position", "Trades", "Wins", "Losses", "Win Rate", "Gross P&L",
              "Net P&L", "Avg Net P&L / Trade", "Best Trade", "Worst Trade"]


def build_position_stats(ws) -> None:
    set_title_subtitle(
        ws, "Position Type Stats",
        "Long vs. short performance, computed straight from Trade Log.",
        len(PS_HEADERS),
    )

    for col, header in enumerate(PS_HEADERS, start=1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(PS_HEADERS))

    widths = [12, 10, 9, 10, 11, 14, 14, 16, 14, 14]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    tl = "'Trade Log'"
    col_fmt = {"B": NUM, "C": NUM, "D": NUM, "E": PCT1, "F": INR, "G": INR,
               "H": INR, "I": INR, "J": INR}

    for r, position in ((4, "LONG"), (5, "SHORT")):
        cell_a = ws[f"A{r}"]; cell_a.value = position; style_formula(cell_a)

        formulas = {
            "B": f"=COUNTIF({tl}!$C$4:$C$503,$A{r})",
            "C": f'=COUNTIFS({tl}!$C$4:$C$503,$A{r},{tl}!$K$4:$K$503,">0")',
            "D": f'=COUNTIFS({tl}!$C$4:$C$503,$A{r},{tl}!$K$4:$K$503,"<0")',
            "E": f'=IFERROR($C{r}/$B{r},"")',
            "F": f"=SUMIF({tl}!$C$4:$C$503,$A{r},{tl}!$I$4:$I$503)",
            "G": f"=SUMIF({tl}!$C$4:$C$503,$A{r},{tl}!$K$4:$K$503)",
            "H": f'=IFERROR($G{r}/$B{r},"")',
            "I": f'=IFERROR(_xlfn.MAXIFS({tl}!$K$4:$K$503,{tl}!$C$4:$C$503,$A{r}),"")',
            "J": f'=IFERROR(_xlfn.MINIFS({tl}!$K$4:$K$503,{tl}!$C$4:$C$503,$A{r}),"")',
        }
        for col_letter, formula in formulas.items():
            cell = ws[f"{col_letter}{r}"]
            cell.value = formula
            style_formula(cell, col_fmt[col_letter])

    # TOTAL row.
    total_row = 6
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = TOTAL_FONT

    total_formulas = {
        "B": "=SUM(B4:B5)",
        "C": "=SUM(C4:C5)",
        "D": "=SUM(D4:D5)",
        "E": f'=IFERROR(C{total_row}/B{total_row},"")',
        "F": "=SUM(F4:F5)",
        "G": "=SUM(G4:G5)",
        "H": f'=IFERROR(G{total_row}/B{total_row},"")',
    }
    for col_letter, formula in total_formulas.items():
        cell = ws[f"{col_letter}{total_row}"]
        cell.value = formula
        cell.font = TOTAL_FONT
        cell.number_format = col_fmt[col_letter]


# ── Sheet 1: Total PnL (dashboard) ──────────────────────────────────────────
# Layout map (all formulas identical in substance to the original plain-list
# version -- only cell *addresses* moved, to fit the card/table/chart grid):
#   Row 1     Title (A1:P1)
#   Row 2     Subtitle (A2:P2)
#   Row 3     Base Capital -- manual input (A3 label, B3 value)
#   Row 4     Band: STRATEGY SNAPSHOT
#   Row 5     KPI card labels
#   Rows 6-8  KPI card numbers (merged 3 rows tall) -- Net P&L ₹ / Net P&L % /
#             Win Rate / Profit Factor, in that column order
#   Row 9     spacer
#   Row 10    Band: DETAILED STATS
#   Row 11    secondary table header (Metric / Value)
#   Rows12-20 secondary stats (Total Trades ... Gross P&L %)
#   Row 21    spacer
#   Row 22    Band: VISUALS
#   Row 23    spacer
#   Row 24+   3 charts, side by side

BC_ROW          = 3
BAND_SNAPSHOT   = 4
CARD_LABEL_ROW  = 5
CARD_NUM_TOP    = 6
CARD_NUM_BOTTOM = 8
BAND_STATS      = 10
STATS_HDR_ROW   = 11
STATS_FIRST_ROW = 12
STATS_LAST_ROW  = 20
BAND_VISUALS    = 22
CHART_ROW       = 24
DASHBOARD_COLS  = 16  # A..P

BAND_FONT = Font(name=FONT_NAME, bold=True, size=11, color="FF1F3864")
BAND_FILL = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

CARD_LABEL_FONT = Font(name=FONT_NAME, size=9, color="FFFFFFFF")
CARD_NUM_FONT   = Font(name=FONT_NAME, size=26, bold=True, color="FFFFFFFF")

GREEN_FILL = PatternFill(start_color="FF2E7D32", end_color="FF2E7D32", fill_type="solid")
RED_FILL   = PatternFill(start_color="FFC62828", end_color="FFC62828", fill_type="solid")

CARD_NAVY   = "FF1F3864"   # Net P&L (₹) base, before conditional flip
CARD_TEAL   = "FF1B7A75"   # Net P&L (%) base, before conditional flip
CARD_AMBER  = "FFC77D02"   # Win Rate, static
CARD_PURPLE = "FF5B2C83"   # Profit Factor, static

# (col_start, col_end, label, formula, number_format, base_fill_hex, conditional)
CARD_DEFS = [
    (2,  4,  "NET P&L (₹)",  "=SUM('Trade Log'!$K$4:$K$503)", INR,   CARD_NAVY,  True),
    (6,  8,  "NET P&L (%)",  '=IF($B$3=0,"",$B$6/$B$3)',      PCT,   CARD_TEAL,  True),
    (10, 12, "WIN RATE",     '=IFERROR($C$13/$C$12,"")',      PCT1,  CARD_AMBER, False),
    (14, 16, "PROFIT FACTOR",
     '=IFERROR(SUMIF(\'Trade Log\'!$K$4:$K$503,">0")/'
     'ABS(SUMIF(\'Trade Log\'!$K$4:$K$503,"<0")),"")', XMULT, CARD_PURPLE, False),
]

# (row, label, formula, number_format, databar)
STATS_ROWS = [
    (12, "Total Trades",      "=COUNT('Trade Log'!$K$4:$K$503)", NUM,  False),
    (13, "Winning Trades",    '=COUNTIF(\'Trade Log\'!$K$4:$K$503,">0")', NUM, False),
    (14, "Losing Trades",     '=COUNTIF(\'Trade Log\'!$K$4:$K$503,"<0")', NUM, False),
    (15, "Average Win (₹)",  '=IFERROR(AVERAGEIF(\'Trade Log\'!$K$4:$K$503,">0"),"")', INR, True),
    (16, "Average Loss (₹)", '=IFERROR(AVERAGEIF(\'Trade Log\'!$K$4:$K$503,"<0"),"")', INR, True),
    (17, "Best Trade (₹)",   "=IFERROR(MAX('Trade Log'!$K$4:$K$503),\"\")", INR, True),
    (18, "Worst Trade (₹)",  "=IFERROR(MIN('Trade Log'!$K$4:$K$503),\"\")", INR, True),
    (19, "Gross P&L (₹)",    "=SUM('Trade Log'!$I$4:$I$503)", INR, False),
    (20, "Gross P&L (%)",    '=IF($B$3=0,"",$C$19/$B$3)', PCT, False),
]


def _band(ws, row: int, text: str) -> None:
    last_col = get_column_letter(DASHBOARD_COLS)
    ws.merge_cells(f"A{row}:{last_col}{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = BAND_FONT
    cell.fill = BAND_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    for col in range(1, DASHBOARD_COLS + 1):
        ws.cell(row=row, column=col).fill = BAND_FILL


def _kpi_card(ws, col_start: int, col_end: int, label: str, formula: str,
              number_format: str, fill_hex: str, conditional: bool) -> str:
    """Builds one KPI card (label strip + big merged number block). Returns
    the number cell's address (e.g. 'B6') so callers can cross-reference it
    and, for conditional cards, anchor the green/red flip rule on it."""
    first_col_letter = get_column_letter(col_start)
    last_col_letter  = get_column_letter(col_end)
    label_range = f"{first_col_letter}{CARD_LABEL_ROW}:{last_col_letter}{CARD_LABEL_ROW}"
    num_range   = f"{first_col_letter}{CARD_NUM_TOP}:{last_col_letter}{CARD_NUM_BOTTOM}"

    ws.merge_cells(label_range)
    ws.merge_cells(num_range)

    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    for row in (CARD_LABEL_ROW, *range(CARD_NUM_TOP, CARD_NUM_BOTTOM + 1)):
        for col in range(col_start, col_end + 1):
            ws.cell(row=row, column=col).fill = fill

    label_cell = ws[f"{first_col_letter}{CARD_LABEL_ROW}"]
    label_cell.value = label
    label_cell.font = CARD_LABEL_FONT
    label_cell.alignment = CENTER

    num_cell = ws[f"{first_col_letter}{CARD_NUM_TOP}"]
    num_cell.value = formula
    num_cell.font = CARD_NUM_FONT
    num_cell.alignment = CENTER
    num_cell.number_format = number_format

    if conditional:
        anchor = f"${first_col_letter}${CARD_NUM_TOP}"
        for rng in (label_range, num_range):
            ws.conditional_formatting.add(rng, FormulaRule(formula=[f"{anchor}>=0"], fill=GREEN_FILL))
            ws.conditional_formatting.add(rng, FormulaRule(formula=[f"{anchor}<0"], fill=RED_FILL))

    return num_cell.coordinate


def build_total_pnl_dashboard(ws, ws_day) -> None:
    set_title_subtitle(
        ws, "Total PnL",
        "Overall strategy performance -- set Base Capital below; everything "
        "else calculates automatically from Trade Log.",
        DASHBOARD_COLS,
    )

    # Column widths -- wide "card area" columns, narrow gap/margin columns.
    ws.column_dimensions["A"].width = 4
    for group_start in (2, 6, 10, 14):  # B, F, J, N
        for offset in range(3):
            ws.column_dimensions[get_column_letter(group_start + offset)].width = 15
    for gap_col in (5, 9, 13):  # E, I, M
        ws.column_dimensions[get_column_letter(gap_col)].width = 8

    # Base Capital -- manual input, unchanged from the original.
    ws["A3"].value = "Base Capital (₹)"
    ws["A3"].font = LABEL_FONT
    style_input(ws["B3"], INR)
    ws["B3"].value = 1_500_000

    # Section 1: KPI cards.
    _band(ws, BAND_SNAPSHOT, "STRATEGY SNAPSHOT")
    for col_start, col_end, label, formula, fmt, fill_hex, conditional in CARD_DEFS:
        _kpi_card(ws, col_start, col_end, label, formula, fmt, fill_hex, conditional)

    # Section 2: secondary stats table (plain, compact, 2 columns).
    _band(ws, BAND_STATS, "DETAILED STATS")
    ws["B11"].value = "Metric"
    ws["C11"].value = "Value"
    for addr in ("B11", "C11"):
        ws[addr].font = HEADER_FONT
        ws[addr].fill = HEADER_FILL
        ws[addr].alignment = CENTER

    for row, label, formula, fmt, databar in STATS_ROWS:
        cell_b = ws[f"B{row}"]
        cell_b.value = label
        cell_b.font = LABEL_FONT

        cell_c = ws[f"C{row}"]
        cell_c.value = formula
        style_formula(cell_c, fmt)

    # Data bars on the four currency stats (Average Win/Loss, Best/Worst Trade),
    # scaled together so relative magnitude reads at a glance.
    ws.conditional_formatting.add(
        "C15:C18",
        DataBarRule(start_type="min", start_value=None, end_type="max", end_value=None,
                    color="638EC6"),
    )

    # Section 3: charts.
    _band(ws, BAND_VISUALS, "VISUALS")

    # Chart 1 -- equity curve (Day Wise PnL cumulative net P&L).
    line = LineChart()
    line.title = "Equity Curve"
    line.style = 2
    line.y_axis.title = "Cumulative Net P&L (₹)"
    line.x_axis.title = "Date"
    line.width = 13
    line.height = 8
    data = Reference(ws_day, min_col=8, min_row=3, max_row=DW_LAST_ROW)
    line.add_data(data, titles_from_data=True)
    cats = Reference(ws_day, min_col=1, min_row=DW_FIRST_ROW, max_row=DW_LAST_ROW)
    line.set_categories(cats)
    line.series[0].smooth = True
    line.series[0].graphicalProperties.line.solidFill = "1F3864"
    line.series[0].graphicalProperties.line.width = 20000
    line.y_axis.majorGridlines = ChartLines(
        spPr=GraphicalProperties(ln=LineProperties(solidFill="D9D9D9"))
    )
    ws.add_chart(line, f"B{CHART_ROW}")

    # Chart 2 -- Win / Loss split (pie), pulling live from the KPI table cells.
    pie = PieChart()
    pie.title = "Win / Loss Split"
    pie.width = 13
    pie.height = 8
    pie_data = Reference(ws, min_col=3, min_row=13, max_row=14)  # C13:C14
    pie.add_data(pie_data, titles_from_data=False)
    pie_cats = Reference(ws, min_col=2, min_row=13, max_row=14)  # B13:B14
    pie.set_categories(pie_cats)
    pie.series[0].data_points = [
        DataPoint(idx=0, spPr=GraphicalProperties(solidFill="2E7D32")),  # wins -- green
        DataPoint(idx=1, spPr=GraphicalProperties(solidFill="C62828")),  # losses -- red
    ]
    ws.add_chart(pie, f"H{CHART_ROW}")

    # Chart 3 -- Gross vs Net P&L (bar), from this sheet's own cells (C19, B6).
    bar = BarChart()
    bar.type = "col"
    bar.title = "Gross vs Net P&L"
    bar.width = 13
    bar.height = 8
    bar.add_data(Reference(ws, min_col=3, min_row=19, max_row=19), titles_from_data=False)  # Gross
    bar.add_data(Reference(ws, min_col=2, min_row=6,  max_row=6),  titles_from_data=False)  # Net
    bar.series[0].tx = SeriesLabel(v="Gross P&L")
    bar.series[1].tx = SeriesLabel(v="Net P&L")
    bar.series[0].graphicalProperties.solidFill = "4472C4"
    bar.series[1].graphicalProperties.solidFill = "2E7D32"
    ws.add_chart(bar, f"N{CHART_ROW}")

    # Keep title/subtitle/Base Capital/band/cards visible while scrolling.
    ws.freeze_panes = "A9"


# ── Build ─────────────────────────────────────────────────────────────────────

def main() -> None:
    # Read state from the OLD file (if any) before it gets overwritten below.
    existing_costs = _load_existing_costs(OUT_PATH)
    trades = load_trades_from_positions()

    wb = Workbook()

    ws_total = wb.active
    ws_total.title = "Total PnL"
    ws_log   = wb.create_sheet("Trade Log")
    ws_day   = wb.create_sheet("Day Wise PnL")
    ws_stats = wb.create_sheet("Position Type Stats")

    build_trade_log(ws_log, trades, existing_costs)
    build_day_wise(ws_day)
    build_total_pnl_dashboard(ws_total, ws_day)
    build_position_stats(ws_stats)

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH} ({len(trades)} synced trade(s) from Dhan positions, "
          f"from {PNL_START_DATE} onward)" if trades else f"Wrote {OUT_PATH} (no synced trades yet -- examples shown)")


if __name__ == "__main__":
    main()
