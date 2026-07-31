"""
zerodha/build_trade_book_xlsx.py — Visual, day-boxed Excel trade book
========================================================================
Reads   : results/trade_book.csv   (written by build_trade_book.py)
Writes  : results/trade_book.xlsx

Two sheets:
  Dashboard  — KPI tiles (win rate, gross/net PnL, charges) + a daily
               Net PnL bar chart, colored green/red per day.
  Trade Book — trades grouped into a bordered "box" per entry day, each
               row tagged 🟢 WIN / 🔴 LOSS / ⏳ OPEN and colored to match,
               with in-cell data bars on the PnL columns.

Usage:
    python zerodha/build_trade_book_xlsx.py
"""

import csv
import sys
from pathlib import Path
from itertools import groupby

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.shapes import GraphicalProperties

_ROOT     = Path(__file__).resolve().parent.parent
_CSV_FILE = _ROOT / "results" / "trade_book.csv"
_OUT_FILE = _ROOT / "results" / "trade_book.xlsx"

_COLUMNS = [
    "Result", "Stock Name", "Position entry date", "Position Exit date", "No of shares",
    "Entry Price", "Exit Price", "Realised PnL", "Realised PnL Pct",
    "Total Charges", "Net PnL", "Net PnL Pct",
]
_MONEY_COLS = {"Entry Price", "Exit Price", "Realised PnL", "Total Charges", "Net PnL"}
_PCT_COLS   = {"Realised PnL Pct", "Net PnL Pct"}
_INT_COLS   = {"No of shares"}

# ── Palette ──────────────────────────────────────────────────────────────
_GREEN       = PatternFill("solid", fgColor="C6EFCE")
_GREEN_DARK  = "2E7D32"
_RED         = PatternFill("solid", fgColor="FFC7CE")
_RED_DARK    = "C62828"
_AMBER       = PatternFill("solid", fgColor="FFF2CC")
_GREY        = PatternFill("solid", fgColor="F2F2F2")
_NAVY        = PatternFill("solid", fgColor="1F3864")
_BLUE        = PatternFill("solid", fgColor="305496")
_LIGHT_BLUE  = PatternFill("solid", fgColor="D9E1F2")
_WHITE       = PatternFill("solid", fgColor="FFFFFF")

_GREEN_FONT  = Font(color="006100")
_RED_FONT    = Font(color="9C0006")
_AMBER_FONT  = Font(color="7F6000", italic=True)
_WHITE_BOLD  = Font(color="FFFFFF", bold=True, size=13)
_WHITE_TITLE = Font(color="FFFFFF", bold=True, size=22)
_BOLD        = Font(bold=True)
_KPI_LABEL   = Font(color="FFFFFF", size=10, bold=True)
_KPI_VALUE   = Font(color="FFFFFF", size=20, bold=True)

_THIN  = Side(style="thin", color="B7B7B7")
_BOX   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_THICK = Side(style="medium", color="1F3864")
_OUTER = Border(left=_THICK, right=_THICK, top=_THICK, bottom=_THICK)


def _to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _load_rows() -> list:
    if not _CSV_FILE.exists():
        sys.exit(f"[trade_book_xlsx] No trade book: {_CSV_FILE} — run build_trade_book.py first.")
    with open(_CSV_FILE, newline="") as f:
        rows = list(csv.DictReader(f))
    rows.sort(key=lambda r: r["Position entry date"])
    return rows


def _kpi_tile(ws, row, col_start, col_end, label, value, fill):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    ws.merge_cells(start_row=row + 1, start_column=col_start, end_row=row + 2, end_column=col_end)

    top = ws.cell(row=row, column=col_start, value=label)
    top.font, top.fill, top.alignment = _KPI_LABEL, fill, Alignment(horizontal="center")

    val = ws.cell(row=row + 1, column=col_start, value=value)
    val.font, val.fill, val.alignment = _KPI_VALUE, fill, Alignment(horizontal="center", vertical="center")

    for r in (row, row + 1, row + 2):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = _OUTER


def _build_dashboard(wb, rows):
    ws = wb.create_sheet("📊 Dashboard")
    for c, w in zip(range(1, 13), [3] * 12):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.merge_cells("A1:L2")
    title = ws.cell(row=1, column=1, value="📊  TRADING DASHBOARD")
    title.font, title.fill = _WHITE_TITLE, _NAVY
    title.alignment = Alignment(horizontal="center", vertical="center")
    for r in (1, 2):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = _NAVY
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24

    closed = [r for r in rows if r["Position Exit date"]]
    open_ps = [r for r in rows if not r["Position Exit date"]]
    wins    = [r for r in closed if (_to_float(r["Realised PnL"]) or 0) > 0]
    losses  = [r for r in closed if (_to_float(r["Realised PnL"]) or 0) < 0]
    gross   = sum(_to_float(r["Realised PnL"]) or 0 for r in closed)
    charges = sum(_to_float(r["Total Charges"]) or 0 for r in closed if r["Total Charges"])
    net     = sum(_to_float(r["Net PnL"]) or 0 for r in closed if r["Net PnL"] not in (None, ""))
    win_rate = (len(wins) / len(closed) * 100) if closed else 0

    tiles = [
        ("TOTAL TRADES",  f"{len(rows)}",              _BLUE),
        ("CLOSED / OPEN", f"{len(closed)} / {len(open_ps)}", _BLUE),
        ("WIN RATE 🎯",   f"{win_rate:.0f}%  ({len(wins)}W-{len(losses)}L)", _GREEN_DARK if win_rate >= 50 else "BF9000"),
        ("GROSS PnL",     f"₹{gross:,.0f}",             _GREEN_DARK if gross >= 0 else _RED_DARK),
        ("TOTAL CHARGES", f"₹{charges:,.0f}",           "7F7F7F"),
        ("NET PnL 💰",    f"₹{net:,.0f}",                _GREEN_DARK if net >= 0 else _RED_DARK),
    ]
    col = 1
    for label, value, color in tiles:
        fill = PatternFill("solid", fgColor=color) if isinstance(color, str) else color
        _kpi_tile(ws, 4, col, col + 1, label, value, fill)
        col += 2
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 8

    # ── Daily summary table (also feeds the chart) ──────────────────────
    header_row = 9
    ws.cell(row=header_row, column=1, value="Entry Date").font = _BOLD
    ws.cell(row=header_row, column=2, value="Trades").font = _BOLD
    ws.cell(row=header_row, column=3, value="Wins").font = _BOLD
    ws.cell(row=header_row, column=4, value="Losses").font = _BOLD
    ws.cell(row=header_row, column=5, value="Gross PnL").font = _BOLD
    ws.cell(row=header_row, column=6, value="Net PnL").font = _BOLD
    for c in range(1, 7):
        ws.cell(row=header_row, column=c).fill = _LIGHT_BLUE
        ws.cell(row=header_row, column=c).border = _BOX

    r = header_row + 1
    first_data_row = r
    for entry_date, day_rows in groupby(rows, key=lambda x: x["Position entry date"]):
        day_rows = list(day_rows)
        day_closed = [x for x in day_rows if x["Position Exit date"]]
        day_wins   = [x for x in day_closed if (_to_float(x["Realised PnL"]) or 0) > 0]
        day_losses = [x for x in day_closed if (_to_float(x["Realised PnL"]) or 0) < 0]
        day_gross  = sum(_to_float(x["Realised PnL"]) or 0 for x in day_closed)
        day_net    = sum(_to_float(x["Net PnL"]) or 0 for x in day_closed if x["Net PnL"] not in (None, ""))

        ws.cell(row=r, column=1, value=entry_date)
        ws.cell(row=r, column=2, value=len(day_rows))
        ws.cell(row=r, column=3, value=len(day_wins))
        ws.cell(row=r, column=4, value=len(day_losses))
        gc = ws.cell(row=r, column=5, value=round(day_gross, 2))
        nc = ws.cell(row=r, column=6, value=round(day_net, 2))
        gc.number_format = nc.number_format = '"₹"#,##0.00'
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = _BOX
        r += 1
    last_data_row = r - 1

    # ── Bar chart: daily Net PnL, colored per bar ───────────────────────
    chart = BarChart()
    chart.title = "Daily Net PnL"
    chart.style = 10
    chart.y_axis.title = "₹ Net PnL"
    chart.x_axis.title = "Entry Date"
    chart.width, chart.height = 26, 11

    data = Reference(ws, min_col=6, min_row=header_row, max_row=last_data_row)
    cats = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None

    series = chart.series[0]
    for i, x in enumerate(range(first_data_row, last_data_row + 1)):
        val = ws.cell(row=x, column=6).value or 0
        color = "2E7D32" if val >= 0 else "C62828"
        pt = DataPoint(idx=i)
        pt.graphicalProperties = GraphicalProperties(solidFill=color)
        series.data_points.append(pt)

    ws.add_chart(chart, f"H{header_row}")
    ws.sheet_view.showGridLines = False
    return ws


def _build_trade_sheet(wb, rows):
    ws = wb.create_sheet("📅 Trade Book")
    ws.sheet_view.showGridLines = False
    n_cols = len(_COLUMNS)
    r = 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    title_cell = ws.cell(row=r, column=1, value="📅  TRADE BOOK — DAY-BY-DAY LOG")
    title_cell.font = Font(bold=True, size=16, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 26
    r += 2

    for entry_date, day_rows in groupby(rows, key=lambda x: x["Position entry date"]):
        day_rows = list(day_rows)
        gross = sum(_to_float(x["Realised PnL"]) or 0 for x in day_rows)
        net   = sum(_to_float(x["Net PnL"]) or 0 for x in day_rows)
        closed_n = sum(1 for x in day_rows if x["Position Exit date"])
        day_emoji = "🟢" if gross > 0 else ("🔴" if gross < 0 else "⏳")

        box_start = r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
        header_text = (f"  {day_emoji}  {entry_date}   ·   {len(day_rows)} trade(s), {closed_n} closed"
                        f"   ·   Gross PnL: ₹{gross:,.2f}   ·   Net PnL: ₹{net:,.2f}")
        cell = ws.cell(row=r, column=1, value=header_text)
        cell.fill, cell.font = _BLUE, _WHITE_BOLD
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22
        r += 1

        for c, col_name in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=col_name)
            cell.fill, cell.font, cell.border = _LIGHT_BLUE, _BOLD, _BOX
            cell.alignment = Alignment(horizontal="center")
        r += 1

        pnl_col_first_row = r
        for row in day_rows:
            realised = _to_float(row["Realised PnL"])
            is_closed = row["Position Exit date"] not in (None, "")

            if is_closed and realised is not None:
                if realised > 0:
                    fill, font, tag = _GREEN, _GREEN_FONT, "🟢 WIN"
                elif realised < 0:
                    fill, font, tag = _RED, _RED_FONT, "🔴 LOSS"
                else:
                    fill, font, tag = _GREY, Font(), "⚪ FLAT"
            else:
                fill, font, tag = _AMBER, _AMBER_FONT, "⏳ OPEN"

            for c, col_name in enumerate(_COLUMNS, start=1):
                if col_name == "Result":
                    val = tag
                else:
                    raw = row.get(col_name, "")
                    val = _to_float(raw) if col_name in (_MONEY_COLS | _PCT_COLS | _INT_COLS) else raw
                    if val == "" or val is None:
                        val = None
                cell = ws.cell(row=r, column=c, value=val)
                cell.fill, cell.font, cell.border = fill, font, _BOX
                if col_name in _MONEY_COLS:
                    cell.number_format = '"₹"#,##0.00'
                elif col_name in _PCT_COLS:
                    cell.number_format = "0.00\"%\""
                elif col_name in _INT_COLS:
                    cell.number_format = "0"
            r += 1
        pnl_col_last_row = r - 1

        # In-cell data bars on Realised PnL & Net PnL, if any closed rows this day
        if pnl_col_last_row >= pnl_col_first_row:
            for col_name, color in (("Realised PnL", "63C384"), ("Net PnL", "63C384")):
                col_idx = _COLUMNS.index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                rng = f"{col_letter}{pnl_col_first_row}:{col_letter}{pnl_col_last_row}"
                ws.conditional_formatting.add(
                    rng, DataBarRule(start_type="min", end_type="max", color=color, showValue=True)
                )

        # Day subtotal row
        ws.cell(row=r, column=1, value="Σ Day total").font = _BOLD
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).fill = _LIGHT_BLUE
            ws.cell(row=r, column=c).border = _BOX
        gc = ws.cell(row=r, column=_COLUMNS.index("Realised PnL") + 1, value=round(gross, 2))
        nc = ws.cell(row=r, column=_COLUMNS.index("Net PnL") + 1, value=round(net, 2))
        gc.font = nc.font = _BOLD
        gc.number_format = nc.number_format = '"₹"#,##0.00'
        box_end = r
        r += 2

        # Thick outer border around the whole day box
        for row_i in range(box_start, box_end + 1):
            ws.cell(row=row_i, column=1).border = Border(
                left=_THICK, right=ws.cell(row=row_i, column=1).border.right,
                top=ws.cell(row=row_i, column=1).border.top, bottom=ws.cell(row=row_i, column=1).border.bottom)
            ws.cell(row=row_i, column=n_cols).border = Border(
                right=_THICK, left=ws.cell(row=row_i, column=n_cols).border.left,
                top=ws.cell(row=row_i, column=n_cols).border.top, bottom=ws.cell(row=row_i, column=n_cols).border.bottom)
        for c in range(1, n_cols + 1):
            top_cell = ws.cell(row=box_start, column=c)
            top_cell.border = Border(top=_THICK, left=top_cell.border.left, right=top_cell.border.right, bottom=top_cell.border.bottom)
            bot_cell = ws.cell(row=box_end, column=c)
            bot_cell.border = Border(bottom=_THICK, left=bot_cell.border.left, right=bot_cell.border.right, top=bot_cell.border.top)

    widths = [10, 13, 15, 15, 10, 11, 11, 12, 13, 12, 11, 11]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"


def main() -> None:
    rows = _load_rows()
    if not rows:
        print("[trade_book_xlsx] No rows — nothing to write.")
        return

    wb = Workbook()
    wb.remove(wb.active)
    _build_dashboard(wb, rows)
    _build_trade_sheet(wb, rows)
    wb.active = 0

    wb.save(_OUT_FILE)
    print(f"[trade_book_xlsx] Wrote {_OUT_FILE}")


if __name__ == "__main__":
    main()
