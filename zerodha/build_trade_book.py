"""
zerodha/build_trade_book.py — Flatten positions_zerodha.json into a P&L-ready trade book
==========================================================================================
Reads   : results/positions_zerodha.json  (live trade state, written by run_trades.py)
Writes  : results/trade_book.csv          (one row per position, git-diffable source of truth)
          results/trade_book.xlsx         (same data, day-boxed visual view)

Columns: Stock Name, Position entry date, Position Exit date, No of shares,
         Entry Price, Exit Price, Realised PnL, Realised PnL Pct,
         Total Charges, Net PnL, Net PnL Pct

Open positions (not yet exited, at any stage -- 945, 1200, or a 945-nodata partial)
get entry-side fields only; exit date/price/P&L/charges stay blank until they close.

Charges are pulled directly from Kite's /charges/orders API (actual, not estimated) for
the entry (BUY) and exit (SELL) order of each closed position, plus a flat ₹15.34 DP
charge on the exit leg (DP charges aren't exposed by any Kite API — confirmed against
Zerodha Console manually).

The xlsx has one sheet, Trade Book: trades grouped into a bordered box per entry
day, each row tagged 🟢 WIN / 🔴 LOSS / ⏳ OPEN and colored to match, with
in-cell data bars on the PnL columns. No gridlines.

Usage:
    python zerodha/build_trade_book.py
"""

import csv
import json
import sys
from pathlib import Path
from itertools import groupby

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))

from zerodha.auth import get_session, BASE_URL

_POS_FILE  = _ROOT / "results" / "positions_zerodha.json"
_CSV_FILE  = _ROOT / "results" / "trade_book.csv"
_XLSX_FILE = _ROOT / "results" / "trade_book.xlsx"

# Real capital (₹1.5L) went live on this date -- earlier positions were test
# trades and are excluded from the trade book, not just old/low-stakes ones.
_TRACKING_START_DATE = "2026-07-27"

_FIELDNAMES = [
    "Stock Name", "Position entry date", "Position Exit date", "No of shares",
    "Entry Price", "Exit Price", "Realised PnL", "Realised PnL Pct",
    "Total Charges", "Net PnL", "Net PnL Pct",
]

_EXIT_STAGE_KEYS = {
    # Pre-2026-08-04 positions recorded exit fields under "945"/"1200" (the
    # stage names before the 9:25am/11:59am schedule change) -- both sets of
    # keys are kept permanently so historical trades still read correctly.
    "exited_945":              ("exit_price_945",  "exit_timestamp_945",  "exit_order_id_945"),
    "exited_1200":             ("exit_price_1200", "exit_timestamp_1200", "exit_order_id_1200"),
    "partial_exit_945_nodata": ("exit_price_945",  "exit_timestamp_945",  "exit_order_id_945"),
    "exited_925":              ("exit_price_925",  "exit_timestamp_925",  "exit_order_id_925"),
    "exited_1159":             ("exit_price_1159", "exit_timestamp_1159", "exit_order_id_1159"),
    "partial_exit_925_nodata": ("exit_price_925",  "exit_timestamp_925",  "exit_order_id_925"),
}

DP_CHARGE = 15.34  # flat, per scrip, sell-side only — not available via any Kite API


# ── Charges (Kite API) ───────────────────────────────────────────────────

def _charge_leg(order_id, symbol, transaction_type, qty, price, product="CNC"):
    return {
        "order_id": order_id, "exchange": "NSE", "tradingsymbol": symbol,
        "transaction_type": transaction_type, "variety": "regular", "product": product,
        "order_type": "MARKET", "quantity": qty, "average_price": price,
    }


def _fetch_charge(leg: dict) -> float:
    """Fetch actual charges for a single order leg.

    IMPORTANT: must be called with exactly one leg per request. Kite's
    /charges/orders silently nets BUY+SELL legs of the *same symbol* within
    a single request (as if it were a same-day intraday square-off), even
    when the two orders are from entirely different calendar days/trades.
    Batching legs together silently produces wrong (understated) charges.
    """
    s, _ = get_session()
    r = s.post(BASE_URL + "/charges/orders", json=[leg])
    r.raise_for_status()
    return r.json()["data"][0]["charges"]["total"]


def _build_rows(positions: list) -> list:
    charges = {}  # position_index -> {"entry": total, "exit": total}
    for i, p in enumerate(positions):
        product   = p.get("product", "CNC")
        entry_leg = _charge_leg(p["entry_order_id"], p["symbol"], "BUY",
                                 p["actual_fill_quantity"], p["actual_fill_price"], product)
        try:
            charges.setdefault(i, {})["entry"] = _fetch_charge(entry_leg)
        except Exception as exc:
            print(f"[trade_book] WARNING: entry charge fetch failed for {p['symbol']} ({exc}).")
            charges.setdefault(i, {})["entry"] = None

        stage_info = _EXIT_STAGE_KEYS.get(p.get("status"))
        if stage_info:
            price_key, _, order_id_key = stage_info
            exit_leg = _charge_leg(p[order_id_key], p["symbol"], "SELL",
                                    p["actual_fill_quantity"], p[price_key], product)
            try:
                charges.setdefault(i, {})["exit"] = _fetch_charge(exit_leg)
            except Exception as exc:
                print(f"[trade_book] WARNING: exit charge fetch failed for {p['symbol']} ({exc}).")
                charges.setdefault(i, {})["exit"] = None

    rows = []
    for i, p in enumerate(positions):
        entry_price = p.get("actual_fill_price")
        qty         = p.get("actual_fill_quantity")

        status = p.get("status")
        exit_price = exit_ts = None
        stage_info = _EXIT_STAGE_KEYS.get(status)
        if stage_info:
            price_key, ts_key, _ = stage_info
            exit_price = p.get(price_key)
            exit_ts    = p.get(ts_key)

        exit_date  = exit_ts.split("T")[0] if exit_ts else None
        realized   = p.get("realized_pnl")
        return_pct = p.get("realized_return_pct")

        total_charges = net_pnl = net_pnl_pct = None
        leg_totals = charges.get(i, {})
        if exit_date and leg_totals.get("entry") is not None and leg_totals.get("exit") is not None:
            total_charges = round(leg_totals["entry"] + leg_totals["exit"] + DP_CHARGE, 2)
            net_pnl       = round(realized - total_charges, 2)
            invested      = entry_price * qty
            net_pnl_pct   = round(net_pnl / invested * 100, 4) if invested else 0

        rows.append({
            "Stock Name":           p.get("symbol"),
            "Position entry date":  p.get("entry_date"),
            "Position Exit date":   exit_date,
            "No of shares":         qty,
            "Entry Price":          entry_price,
            "Exit Price":           exit_price,
            "Realised PnL":         realized,
            "Realised PnL Pct":     return_pct,
            "Total Charges":        total_charges,
            "Net PnL":              net_pnl,
            "Net PnL Pct":          net_pnl_pct,
        })
    return rows


def _write_csv(rows: list) -> None:
    with open(_CSV_FILE, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=_FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)


# ── XLSX (dashboard + day-boxed trade log) ───────────────────────────────

_XLSX_COLUMNS = ["Result"] + _FIELDNAMES
_MONEY_COLS = {"Entry Price", "Exit Price", "Realised PnL", "Total Charges", "Net PnL"}
_PCT_COLS   = {"Realised PnL Pct", "Net PnL Pct"}
_INT_COLS   = {"No of shares"}

_GREEN       = PatternFill("solid", fgColor="C6EFCE")
_RED         = PatternFill("solid", fgColor="FFC7CE")
_AMBER       = PatternFill("solid", fgColor="FFF2CC")
_GREY        = PatternFill("solid", fgColor="F2F2F2")
_BLUE        = PatternFill("solid", fgColor="305496")
_LIGHT_BLUE  = PatternFill("solid", fgColor="D9E1F2")

_GREEN_FONT  = Font(color="006100")
_RED_FONT    = Font(color="9C0006")
_AMBER_FONT  = Font(color="7F6000", italic=True)
_WHITE_BOLD  = Font(color="FFFFFF", bold=True, size=13)
_BOLD        = Font(bold=True)

_THIN  = Side(style="thin", color="B7B7B7")
_BOX   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_THICK = Side(style="medium", color="1F3864")


def _to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _build_trade_sheet(wb, rows):
    ws = wb.create_sheet("📅 Trade Book")
    ws.sheet_view.showGridLines = False
    n_cols = len(_XLSX_COLUMNS)
    r = 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    title_cell = ws.cell(row=r, column=1, value="📅  TRADE BOOK — DAY-BY-DAY LOG")
    title_cell.font = Font(bold=True, size=16, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 26
    r += 2

    for entry_date, day_rows in groupby(rows, key=lambda x: x["Position entry date"]):
        day_rows = list(day_rows)
        gross = sum(_to_float(x["Realised PnL"]) or 0 for x in day_rows)
        net   = sum(_to_float(x["Net PnL"]) or 0 for x in day_rows)
        closed_n = sum(1 for x in day_rows if x["Position Exit date"])
        day_emoji = "🟢" if gross > 0 else ("🔴" if gross < 0 else "⏳")

        box_start = r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
        header_text = (f"  {day_emoji}  {entry_date}   ·   {len(day_rows)} trade(s), {closed_n} closed"
                        f"   ·   Gross PnL: ₹{gross:,.2f}   ·   Net PnL: ₹{net:,.2f}")
        cell = ws.cell(row=r, column=1, value=header_text)
        cell.fill, cell.font = _BLUE, _WHITE_BOLD
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22
        r += 1

        for c, col_name in enumerate(_XLSX_COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=col_name)
            cell.fill, cell.font, cell.border = _LIGHT_BLUE, _BOLD, _BOX
            cell.alignment = Alignment(horizontal="center")
        r += 1

        pnl_col_first_row = r
        for row in day_rows:
            realised = _to_float(row["Realised PnL"])
            is_closed = row["Position Exit date"] not in (None, "")

            if is_closed and realised is not None:
                if realised > 0:
                    fill, font, tag = _GREEN, _GREEN_FONT, "🟢 WIN"
                elif realised < 0:
                    fill, font, tag = _RED, _RED_FONT, "🔴 LOSS"
                else:
                    fill, font, tag = _GREY, Font(), "⚪ FLAT"
            else:
                fill, font, tag = _AMBER, _AMBER_FONT, "⏳ OPEN"

            for c, col_name in enumerate(_XLSX_COLUMNS, start=1):
                if col_name == "Result":
                    val = tag
                else:
                    raw = row.get(col_name, "")
                    val = _to_float(raw) if col_name in (_MONEY_COLS | _PCT_COLS | _INT_COLS) else raw
                    if val == "" or val is None:
                        val = None
                cell = ws.cell(row=r, column=c, value=val)
                cell.fill, cell.font, cell.border = fill, font, _BOX
                if col_name in _MONEY_COLS:
                    cell.number_format = '"₹"#,##0.00'
                elif col_name in _PCT_COLS:
                    cell.number_format = "0.00\"%\""
                elif col_name in _INT_COLS:
                    cell.number_format = "0"
            r += 1
        pnl_col_last_row = r - 1

        # In-cell data bars on Realised PnL & Net PnL, if any closed rows this day
        if pnl_col_last_row >= pnl_col_first_row:
            for col_name in ("Realised PnL", "Net PnL"):
                col_idx = _XLSX_COLUMNS.index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                rng = f"{col_letter}{pnl_col_first_row}:{col_letter}{pnl_col_last_row}"
                ws.conditional_formatting.add(
                    rng, DataBarRule(start_type="min", end_type="max", color="63C384", showValue=True)
                )

        # Day subtotal row
        ws.cell(row=r, column=1, value="Σ Day total").font = _BOLD
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).fill = _LIGHT_BLUE
            ws.cell(row=r, column=c).border = _BOX
        gc = ws.cell(row=r, column=_XLSX_COLUMNS.index("Realised PnL") + 1, value=round(gross, 2))
        nc = ws.cell(row=r, column=_XLSX_COLUMNS.index("Net PnL") + 1, value=round(net, 2))
        gc.font = nc.font = _BOLD
        gc.number_format = nc.number_format = '"₹"#,##0.00'
        box_end = r
        r += 2

        # Thick outer border around the whole day box
        for row_i in range(box_start, box_end + 1):
            left_cell = ws.cell(row=row_i, column=1)
            left_cell.border = Border(left=_THICK, right=left_cell.border.right,
                                       top=left_cell.border.top, bottom=left_cell.border.bottom)
            right_cell = ws.cell(row=row_i, column=n_cols)
            right_cell.border = Border(right=_THICK, left=right_cell.border.left,
                                        top=right_cell.border.top, bottom=right_cell.border.bottom)
        for c in range(1, n_cols + 1):
            top_cell = ws.cell(row=box_start, column=c)
            top_cell.border = Border(top=_THICK, left=top_cell.border.left,
                                      right=top_cell.border.right, bottom=top_cell.border.bottom)
            bot_cell = ws.cell(row=box_end, column=c)
            bot_cell.border = Border(bottom=_THICK, left=bot_cell.border.left,
                                      right=bot_cell.border.right, top=bot_cell.border.top)

    widths = [10, 13, 15, 15, 10, 11, 11, 12, 13, 12, 11, 11]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _write_xlsx(rows: list) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _build_trade_sheet(wb, rows)
    wb.active = 0
    wb.save(_XLSX_FILE)


# ── Main ──────────────────────────────────────────────────────────────────

def main() -> None:
    if not _POS_FILE.exists():
        sys.exit(f"[trade_book] No positions file: {_POS_FILE}")

    positions = json.loads(_POS_FILE.read_text())
    n_before = len(positions)
    positions = [p for p in positions
                 if (p.get("entry_date") or "") >= _TRACKING_START_DATE]
    if n_before != len(positions):
        print(f"[trade_book] Excluded {n_before - len(positions)} position(s) "
              f"before {_TRACKING_START_DATE} (pre-live-capital test trades).")
    if not positions:
        print("[trade_book] No positions on record — nothing to write.")
        return

    rows = _build_rows(positions)

    _write_csv(rows)
    print(f"[trade_book] Wrote {len(rows)} rows -> {_CSV_FILE}")

    _write_xlsx(rows)
    print(f"[trade_book] Wrote {_XLSX_FILE}")

    closed     = [r for r in rows if r["Realised PnL"] is not None]
    open_ps    = [r for r in rows if r["Position Exit date"] is None]
    realized   = sum(r["Realised PnL"] for r in closed)
    net_closed = [r for r in closed if r["Net PnL"] is not None]
    net        = sum(r["Net PnL"] for r in net_closed)

    print(f"[trade_book] Closed: {len(closed)}  Realized P&L: {round(realized, 2):+,.2f}")
    print(f"[trade_book] Open: {len(open_ps)}")
    if net_closed:
        print(f"[trade_book] Net P&L (after charges, {len(net_closed)} rows): {round(net, 2):+,.2f}")


if __name__ == "__main__":
    main()
